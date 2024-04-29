import openpyxl
import re

# Открыть файл Excel
workbook = openpyxl.load_workbook('111.xlsx')

# Получить активный лист
sheet = workbook.active

# Создать новый лист или очистить существующий, если таковой уже есть
sheet_name = 'Специалисты'
if sheet_name in workbook.sheetnames:
    new_sheet = workbook[sheet_name]
    new_sheet.delete_rows(1, new_sheet.max_row)
else:
    new_sheet = workbook.create_sheet(sheet_name)

# Шаблон для разделения содержимого ячеек
delimiter_pattern = r',|\.|:'

# Множество для хранения уникальных специальностей и исследований
unique_entries = set()

# Обрабатываем третий столбец на первом листе
for cell in sheet['C']:
    items = re.split(delimiter_pattern, cell.value.strip()) if cell.value else []
    for item in items:
        cleaned = item.strip()
        if cleaned:
            unique_entries.add(cleaned)

# Запись уникальных специальностей и исследований на новый лист
for i, entry in enumerate(sorted(unique_entries), start=1):
    new_sheet.cell(row=i, column=1).value = entry

# Сохранить изменения в файле Excel
workbook.save('updated_data.xlsx')

# Вывод сообщения об успешном выполнении операции
print("Файл 'updated_data.xlsx' успешно обновлен и содержит список всех специальностей и исследований.")

########
# Открыть файл Excel
workbook = openpyxl.load_workbook('updated_data.xlsx')

# Получить активный лист с таблицей
main_sheet = workbook.active

# Получить лист со специалистами
specialists_sheet = workbook['Специалисты']

# Список специальных фраз
special_phrases = ["гинеколог", "Мазок на флору", "Мазок на онкоцитологию", "УЗИ органов малого таза"]

# Проходим по каждой специальности на листе со специалистами
for i in range(1, specialists_sheet.max_row + 1):
    specialist = specialists_sheet.cell(row=i, column=1).value
    sum_values = 0  # Переменная для суммирования значений в 5 и 6 столбцах

    # Проверяем, содержит ли специальность одну из особых фраз
    check_special = any(phrase in specialist for phrase in special_phrases)

    # Ищем вхождения на листе с таблицей
    for row in main_sheet.iter_rows(min_row=1):
        if row[2].value and specialist in str(row[2].value):  # Проверяем вхождение специалиста
            # Если специальность содержит одну из особых фраз, учитываем только столбец 6
            if check_special:
                value2 = row[5].value if row[5].value is not None and isinstance(row[5].value, (int, float)) else 0
                sum_values += value2
            else:
                # Иначе учитываем значения из обоих столбцов 5 и 6
                value1 = row[4].value if row[4].value is not None and isinstance(row[4].value, (int, float)) else 0
                value2 = row[5].value if row[5].value is not None and isinstance(row[5].value, (int, float)) else 0

                sum_values += value1 + value2

    # Записываем полученную сумму на листе со специалистами во вторую колонку
    specialists_sheet.cell(row=i, column=2).value = sum_values

# Сохраняем изменения
workbook.save('updated_data_with_special_conditions.xlsx')

# Вывод сообщения об успешном выполнении операции
print("Файл 'updated_data_with_special_conditions.xlsx' успешно обновлен с суммированными значениями.")